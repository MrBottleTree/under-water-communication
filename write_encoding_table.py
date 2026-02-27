#!/usr/bin/env python3
import argparse
import csv
import os
import tempfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


def normalize_character(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if text == " " or text.upper() == "<SPACE>" or text.upper() == "SPACE":
        return "SPACE"
    if text == "\t" or text.upper() == "<TAB>" or text.upper() == "TAB" or text == "\\t":
        return "TAB"
    if text == "\n" or text.upper() == "<NEWLINE>" or text.upper() == "NEWLINE" or text == "\\n":
        return "NEWLINE"
    if text == "\r" or text.upper() == "<CR>" or text.upper() == "CR" or text == "\\r":
        return "CR"
    return text


def find_column_name(columns: Iterable[str], wanted: str) -> Optional[str]:
    wanted_lower = wanted.lower()
    for col in columns:
        if col.strip().lower() == wanted_lower:
            return col
    return None


def read_source_rows(path: Path) -> Tuple[List[str], List[Dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError(f"No header row found in: {path}")
            rows: List[Dict[str, object]] = list(reader)
            return list(reader.fieldnames), rows

    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "Reading .xlsx/.xlsm requires openpyxl. Install it with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            raise ValueError(f"No header row found in: {path}")
        headers = ["" if h is None else str(h) for h in header_cells]
        rows: List[Dict[str, object]] = []
        for vals in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: vals[i] for i in range(len(headers))}
            rows.append(row)
        return headers, rows

    raise ValueError(f"Unsupported input format: {path.suffix}. Use .csv, .xlsx, or .xlsm")


def build_code_map(input_path: Path) -> Dict[str, str]:
    columns, rows = read_source_rows(input_path)
    character_col = find_column_name(columns, "Character")
    code_col = find_column_name(columns, "Code")
    if not character_col or not code_col:
        raise ValueError(
            f"Input file must contain 'Character' and 'Code' columns. Found: {columns}"
        )

    code_map: Dict[str, str] = {}
    for row in rows:
        character = normalize_character(row.get(character_col))
        if character == "":
            continue
        code_value = row.get(code_col)
        code = "" if code_value is None else str(code_value)
        code_map[character] = code
    return code_map


def read_global_table_rows(table_path: Path) -> List[Tuple[str, str]]:
    lines = table_path.read_text(encoding="utf-8-sig").splitlines()
    if not lines:
        raise ValueError(f"No content found in encoding table: {table_path}")

    header = lines[0].strip()
    if header.lower() != "character,code":
        raise ValueError(
            f"Encoding table must start with 'Character,Code'. Found header: {lines[0]}"
        )

    rows: List[Tuple[str, str]] = []
    for line in lines[1:]:
        if "," not in line:
            raise ValueError(f"Invalid encoding table row (missing comma): {line!r}")
        character, _, code = line.rpartition(",")
        rows.append((character, code))
    return rows


def write_global_table_rows(output_path: Path, rows: List[Tuple[str, str]]) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as f:
        f.write("Character,Code\n")
        for character, code in rows:
            f.write(f"{character},{code}\n")


def update_encoding_table(table_path: Path, code_map: Dict[str, str], output_path: Path) -> None:
    rows = read_global_table_rows(table_path)
    updated_rows: List[Tuple[str, str]] = []
    for character, _ in rows:
        normalized_character = normalize_character(character)
        updated_rows.append((character, code_map.get(normalized_character, "")))

    if output_path.resolve() == table_path.resolve():
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="", delete=False) as tmp:
            temp_path = Path(tmp.name)
        try:
            write_global_table_rows(temp_path, updated_rows)
            os.replace(temp_path, table_path)
        finally:
            if temp_path.exists():
                temp_path.unlink()
    else:
        write_global_table_rows(output_path, updated_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fill the Code column of the global encoding table from an input CSV/XLSX file "
            "using Character and Code columns."
        )
    )
    parser.add_argument("input_file", help="Input CSV/XLSX/XLSM file containing Character and Code columns")
    parser.add_argument(
        "--encoding-table",
        default="app/src/main/assets/encoding_table.csv",
        help="Path to the global encoding_table.csv (default: app/src/main/assets/encoding_table.csv)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output CSV path. If omitted, the encoding table is overwritten in place.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_file)
    table_path = Path(args.encoding_table)
    output_path = Path(args.output) if args.output else table_path

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not table_path.exists():
        raise FileNotFoundError(f"Encoding table not found: {table_path}")

    code_map = build_code_map(input_path)
    update_encoding_table(table_path, code_map, output_path)
    print(f"Updated encoding table written to: {output_path}")


if __name__ == "__main__":
    main()
